from flask import Flask, render_template, request, send_file, jsonify
import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os
import io
import json
import re
import time
import random
from urllib.parse import urljoin, urlparse, quote_plus

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

FULL_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.5',
    
    'Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1',
}

FIELD_SELECTORS = {
    'Product Title': [
        'h1', '[class*="product-title"]', '[class*="product-name"]',
        '[class*="pdp-title"]', '[itemprop="name"]', '#productTitle',
        '[class*="product_title"]', '[class*="item-title"]', 'h1.title',
        '.product__title', '[class*="product-single__title"]'
    ],
    'Description': [
        '[class*="product-block-list__item--description"]',
        '[class*="expandable-content"]',
        '[itemprop="description"]',
        '[class*="product-description"]',
        '[class*="product__description"]',
        '#productDescription',
        '[class*="product_description"]',
        '[class*="pdp-description"]',
        '[class*="product-detail"]',
        '[class*="overview"]',
        '#description',
        '[class*="description"]',
        '[class*="desc"]',
    ],
    'Bullet Points / Features': [
        '[class*="key-product-feature"]',
        '[class*="feature-list"]',
        'ul[class*="feature"]',
        'ul[class*="bullet"]',
        'ul[class*="highlight"]',
        '[class*="specs"] ul',
        '[class*="benefits"] ul',
        '[class*="selling-point"]',
        '[class*="usp"]',
        '.product__description ul',
        '.rte ul',
    ],
    'Dimensions & Weight': [
        '[class*="dimension"]', '[class*="weight"]', '[class*="specification"]',
        '[class*="spec"]', '[class*="technical"]', 'table[class*="spec"]',
        '[class*="product-spec"]', '[class*="measurements"]', '[class*="size"]'
    ],
    'Price': [
        '[class*="price"]', '[itemprop="price"]', '[class*="product-price"]',
        '.price', '#price', '[class*="sale-price"]',
        '.product__price', '[class*="product-single__price"]'
    ],
    'Images': [
        '[class*="product-image"] img', '[class*="gallery"] img',
        '[class*="pdp-image"] img', '#product-image img',
        '[class*="main-image"] img', '[itemprop="image"]',
        '.product__photo img', '[class*="product-single__photo"] img'
    ],
    'Washing Instructions': [
        '[class*="wash"]', '[class*="care"]', '[class*="laundry"]',
        '[class*="care-instruction"]', '[class*="cleaning"]'
    ],
    'Materials / Fabric': [
        '[class*="material"]', '[class*="fabric"]', '[class*="composition"]',
        '[class*="ingredient"]', '[class*="content"]'
    ],
}

def normalize_url(url):
    if not url:
        return None
    url = url.strip()
    if not url.startswith('http'):
        url = 'https://' + url
    return url

def find_product_url(brand_website, part_number):
    """Find product page URL from brand website + part number"""
    brand_url = normalize_url(brand_website)
    brand_domain = urlparse(brand_url).netloc

    # Search URL strategies — Shopify, WooCommerce, Magento, generic
    strategies = [
        f"{brand_url}/search?type=product&options%5Bprefix%5D=last&q={quote_plus(part_number)}",
        f"{brand_url}/search?type=product&q={quote_plus(part_number)}",
        f"{brand_url}/search?q={quote_plus(part_number)}",
        f"{brand_url}/?s={quote_plus(part_number)}&post_type=product",
        f"{brand_url}/?s={quote_plus(part_number)}",
        f"{brand_url}/catalogsearch/result/?q={quote_plus(part_number)}",
        f"{brand_url}/search?query={quote_plus(part_number)}",
        f"{brand_url}/search?keywords={quote_plus(part_number)}",
        f"{brand_url}/search?search={quote_plus(part_number)}",
        f"{brand_url}/search?text={quote_plus(part_number)}",
    ]

    # Direct URL guesses (skip search entirely if lucky)
    direct_guesses = [
        f"{brand_url}/products/{part_number.lower()}",
        f"{brand_url}/product/{part_number.lower()}",
        f"{brand_url}/p/{part_number.lower()}",
        f"{brand_url}/items/{part_number.lower()}",
    ]

    # Try direct guesses first
    for guess_url in direct_guesses:
        try:
            resp = requests.get(guess_url, headers=FULL_HEADERS, timeout=8, allow_redirects=True)
            if resp.status_code == 200 and part_number.lower() in resp.text.lower():
                return guess_url
        except Exception:
            continue

    # Try search pages
    for strategy_url in strategies:
        try:
            time.sleep(random.uniform(0.8, 2.0))
            resp = requests.get(strategy_url, headers=FULL_HEADERS, timeout=12, allow_redirects=True)
            if resp.status_code != 200:
                continue

            soup = BeautifulSoup(resp.text, 'lxml')
            links = soup.find_all('a', href=True)
            if not links:
                continue

            # PRIORITY 1: Link text or href contains exact part number
            for link in links:
                href = link.get('href', '').strip()
                text = link.get_text(strip=True)
                if (part_number.lower() in text.lower() or
                        part_number.lower() in href.lower()):
                    full_url = urljoin(brand_url, href)
                    if (brand_domain in full_url and
                            full_url.rstrip('/') != strategy_url.rstrip('/') and
                            not any(kw in full_url.lower() for kw in
                                    ['search?', 'category', '?s=', 'catalogsearch', 'cart', 'login', 'account'])):
                        return full_url

            # PRIORITY 2: First product-pattern link
            product_keywords = ['/products/', '/product/', '/p/', '/item/', '/pdp/']
            seen = set()
            for link in links:
                href = link.get('href', '').strip()
                if any(kw in href.lower() for kw in product_keywords):
                    full_url = urljoin(brand_url, href)
                    clean = full_url.split('?')[0]
                    if brand_domain in full_url and clean not in seen:
                        seen.add(clean)
                        return full_url

        except Exception:
            continue

    return None

def calculate_confidence(value, field_name, part_number, page_text):
    if not value or value.strip() == '':
        return 0, 'Not Found'
    score = 50
    value_lower = value.lower()
    if field_name == 'Product Title':
        if 10 <= len(value) <= 200: score += 20
        if part_number.lower() in value_lower: score += 25
    elif field_name == 'Description':
        if len(value) > 100: score += 20
        if len(value) > 500: score += 10
    elif field_name == 'Bullet Points / Features':
        if '\n' in value or len(value.split('•')) > 2: score += 25
    elif field_name == 'Price':
        if any(c in value for c in ['$', '€', '£', '¥']): score += 30
        if re.search(r'\d+\.\d{2}', value): score += 20
    elif field_name == 'Images':
        if value.startswith('http') and any(ext in value.lower() for ext in ['.jpg', '.jpeg', '.png', '.webp']):
            score += 40
    elif field_name == 'Dimensions & Weight':
        if any(unit in value_lower for unit in ['inch', 'cm', 'mm', 'lb', 'kg', 'oz', '"', ' x ']):
            score += 35
    if len(value) < 3: score -= 30
    if value.count('{') > 2 or 'function' in value: score -= 40
    if value.lower() in ['none', 'null', 'undefined', 'n/a']: score -= 40
    score = max(0, min(100, score))
    if score >= 80: label = '✅ High'
    elif score >= 50: label = '⚠️ Medium'
    else: label = '❌ Low'
    return score, label

def scrape_field(soup, field_name, page_url=None):
    """Smart field scraper — works on any website structure"""

    # ── IMAGES ────────────────────────────────────────────────────────────
    if field_name == 'Images':
        images = []
        selectors = FIELD_SELECTORS.get('Images', [])
        for selector in selectors:
            try:
                elements = soup.select(selector)
                for el in elements[:5]:
                    src = el.get('src') or el.get('data-src') or el.get('data-lazy-src')
                    if src and not src.startswith('data:'):
                        if not src.startswith('http') and page_url:
                            src = urljoin(page_url, src)
                        images.append(src)
            except Exception:
                continue
        return ' | '.join(images[:3]) if images else ''

    # ── PRODUCT TITLE ─────────────────────────────────────────────────────
    if field_name == 'Product Title':
        for selector in FIELD_SELECTORS.get('Product Title', []):
            try:
                el = soup.select_one(selector)
                if el:
                    text = el.get_text(strip=True)
                    if text and len(text) > 3:
                        return text[:500]
            except Exception:
                continue
        return ''

    # ── DESCRIPTION ───────────────────────────────────────────────────────
    # Smart approach: collect ONLY paragraph text, stop before bullet lists
    if field_name == 'Description':
        # Try specific selectors first
        specific_selectors = [
            '[class*="product-block-list__item--description"]',
            '[class*="expandable-content"]',
            '[itemprop="description"]',
            '[class*="product-description"]',
            '[class*="product__description"]',
            '#productDescription',
            '#description',
        ]
        for selector in specific_selectors:
            try:
                el = soup.select_one(selector)
                if el:
                    # Get ONLY paragraph text (not list items)
                    paras = el.find_all('p')
                    if paras:
                        text = ' '.join([p.get_text(strip=True) for p in paras
                                        if len(p.get_text(strip=True)) > 20])
                        if text and len(text) > 50:
                            return text[:2000]
                    # Fallback: get element text but stop before any ul
                    ul = el.find('ul')
                    if ul:
                        # Get text up to the first ul
                        raw = el.get_text(separator='\n', strip=True)
                        ul_text = ul.get_text(strip=True)
                        if ul_text in raw:
                            text = raw[:raw.find(ul_text)].strip()
                        else:
                            text = raw
                    else:
                        text = el.get_text(separator=' ', strip=True)
                    # Remove leading label
                    for label in ['Description', 'Product Description', 'About']:
                        if text.lower().startswith(label.lower()):
                            text = text[len(label):].strip()
                    if text and len(text) > 50:
                        return text[:2000]
            except Exception:
                continue

        # Smart fallback: collect paragraphs with real product content
        try:
            paragraphs = soup.find_all('p')
            good_paras = []
            for p in paragraphs:
                text = p.get_text(strip=True)
                skip_words = ['cookie', 'privacy policy', 'shipping policy', 'return policy',
                             'copyright', 'payment', 'subscribe', 'newsletter', 'sign up',
                             'email', 'follow us', 'free shipping on orders']
                if (len(text) > 80 and
                    not any(skip in text.lower() for skip in skip_words)):
                    good_paras.append(text)
                    if len(good_paras) >= 3:
                        break
            if good_paras:
                return ' '.join(good_paras)[:2000]
        except Exception:
            pass
        return ''

    # ── BULLET POINTS / FEATURES ─────────────────────────────────────────
    # Smart approach: find ANY ul with substantive list items
    if field_name == 'Bullet Points / Features':

        def extract_good_bullets(ul_element):
            """Extract clean bullets from a ul, skip nav/breadcrumb items"""
            lis = ul_element.find_all('li', recursive=False) or ul_element.find_all('li')
            good = []
            for li in lis:
                text = li.get_text(strip=True)
                # Skip short items, breadcrumbs, pure links
                if (len(text) > 25 and
                    text.lower() not in ['home', 'all products', 'shop', 'products', 'sale'] and
                    not (len(li.find_all('a')) > 0 and len(text) < 60)):
                    good.append(text)
            return good

        # STRATEGY 1: Look for heading keywords then find sibling/child ul
        feature_heading_keywords = [
            'feature', 'highlight', 'key', 'detail', 'specification',
            'spec', 'included', "whats in", 'benefit', 'why', 'about'
        ]
        for heading in soup.find_all(['h2', 'h3', 'h4', 'strong', 'b']):
            heading_text = heading.get_text(strip=True).lower()
            if any(kw in heading_text for kw in feature_heading_keywords) and len(heading_text) < 60:
                # Check parent and grandparent for ul
                for ancestor in [heading.parent, heading.parent.parent if heading.parent else None]:
                    if ancestor is None:
                        continue
                    for ul in ancestor.find_all('ul'):
                        bullets = extract_good_bullets(ul)
                        if len(bullets) >= 2:
                            return '\n'.join([f'• {b}' for b in bullets[:8]])

        # STRATEGY 2: Try specific CSS selectors
        specific_ul_selectors = [
            '[class*="key-product-feature"]',
            '[class*="feature-list"]',
            'ul[class*="feature"]',
            'ul[class*="bullet"]',
            'ul[class*="highlight"]',
            '[class*="specs"] ul',
            '[class*="benefits"] ul',
            '.product__description ul',
            '.rte ul',
        ]
        for selector in specific_ul_selectors:
            try:
                el = soup.select_one(selector)
                if el:
                    tag = el if el.name == 'ul' else el.find('ul')
                    if tag:
                        bullets = extract_good_bullets(tag)
                        if len(bullets) >= 2:
                            return '\n'.join([f'• {b}' for b in bullets[:8]])
            except Exception:
                continue

        # STRATEGY 3: Find the best ul on the page by quality score
        best_ul = None
        best_score = 0
        nav_skip = ['breadcrumb', 'nav', 'menu', 'header', 'footer', 'social',
                   'share', 'pagination', 'language', 'currency']

        for ul in soup.find_all('ul'):
            ul_class = ' '.join(ul.get('class', []))
            parent_class = ' '.join(ul.parent.get('class', [])) if ul.parent else ''
            all_classes = (ul_class + ' ' + parent_class).lower()

            # Skip navigation lists
            if any(skip in all_classes for skip in nav_skip):
                continue

            bullets = extract_good_bullets(ul)
            if len(bullets) < 2:
                continue

            # Score this ul
            score = len(bullets)
            # Bonus for longer items (product features tend to be descriptive)
            avg_len = sum(len(b) for b in bullets) / len(bullets)
            if avg_len > 50:
                score += 3
            if avg_len > 100:
                score += 3
            # Bonus if items contain | or : (common feature format)
            if any('|' in b or ':' in b for b in bullets):
                score += 5

            if score > best_score:
                best_score = score
                best_ul = bullets

        if best_ul:
            return '\n'.join([f'• {b}' for b in best_ul[:8]])

        return ''

    # ── OTHER FIELDS ──────────────────────────────────────────────────────
    selectors = FIELD_SELECTORS.get(field_name, [])
    for selector in selectors:
        try:
            element = soup.select_one(selector)
            if element:
                text = element.get_text(separator=' ', strip=True)
                if text and len(text) > 5 and text.lower() not in ['view sizing charts', 'size chart', 'add to cart']:
                    return text[:2000]
        except Exception:
            continue

    # Fallback for custom/other fields
    return ''

def scrape_custom_field(soup, custom_field_name):
    field_lower = custom_field_name.lower().strip()
    all_text = soup.get_text(separator='\n')
    lines = all_text.split('\n')
    for i, line in enumerate(lines):
        if field_lower in line.lower() and len(line.strip()) < 100:
            for j in range(i + 1, min(i + 4, len(lines))):
                next_line = lines[j].strip()
                if next_line and len(next_line) > 2:
                    return next_line
    keywords = field_lower.replace(' ', '-').replace('/', '-').split()
    for keyword in keywords:
        if len(keyword) > 3:
            elements = soup.find_all(attrs={'class': re.compile(keyword, re.I)})
            for el in elements:
                text = el.get_text(strip=True)
                if text and len(text) > 3:
                    return text[:1000]
    return ''

def scrape_product(part_number, brand_website, product_url, selected_fields, custom_fields):
    result = {'Part Number': part_number, 'Source URL': '', 'Status': ''}
    all_fields = selected_fields + custom_fields
    for field in all_fields:
        result[field] = ''
        result[f'{field} — Confidence'] = ''
    try:
        # Step 1: Get URL
        if product_url and product_url.strip():
            url = normalize_url(product_url.strip())
        else:
            url = find_product_url(brand_website, part_number)
        if not url:
            result['Status'] = '❌ Product page not found'
            return result
        result['Source URL'] = url

        # Step 2: Fetch page
        time.sleep(random.uniform(1, 2))
        resp = requests.get(url, headers=FULL_HEADERS, timeout=15, allow_redirects=True)
        if resp.status_code != 200:
            result['Status'] = f'❌ HTTP {resp.status_code}'
            return result

        soup = BeautifulSoup(resp.text, 'lxml')
        page_text = soup.get_text()

        # Step 3: Scrape fields
        for field in selected_fields:
            value = scrape_field(soup, field, url)
            score, label = calculate_confidence(value, field, part_number, page_text)
            result[field] = value
            result[f'{field} — Confidence'] = f'{label} ({score}%)'

        # Step 4: Custom fields
        for field in custom_fields:
            if field.strip():
                value = scrape_custom_field(soup, field)
                score, label = calculate_confidence(value, field, part_number, page_text)
                result[field] = value
                result[f'{field} — Confidence'] = f'{label} ({score}%)'

        result['Status'] = '✅ Scraped'
    except requests.exceptions.Timeout:
        result['Status'] = '❌ Timeout — website too slow'
    except requests.exceptions.ConnectionError:
        result['Status'] = '❌ Cannot connect to website'
    except Exception as e:
        result['Status'] = f'❌ Error: {str(e)[:50]}'
    return result

def create_template_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products to Scrape"
    header_fill = PatternFill(start_color="0D1B3E", end_color="0D1B3E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    example_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    example_font = Font(color="666666", italic=True, name="Calibri", size=10)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC')
    )
    headers = ['Part Number', 'Brand Website', 'Product URL (Optional)']
    col_widths = [20, 30, 50]
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 30
    examples = [
        ['EXAMPLE-001  (Remove this row before upload)', 'brandwebsite.com', 'https://brandwebsite.com/product/example-001  |  Leave blank if you do not have the URL'],
    ]
    for row_idx, example in enumerate(examples, 2):
        for col_idx, value in enumerate(example, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill
            cell.font = example_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
        ws.row_dimensions[row_idx].height = 20
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ("Pattern Website Scraper Tool — Template Instructions", True),
        ("", False),
        ("Column 1 — Part Number: Enter the part number provided by the brand.", False),
        ("Column 2 — Brand Website: Just the domain, e.g. petsafe.com", False),
        ("Column 3 — Product URL (Optional): Direct link to product page. Leave blank and the tool will find it!", False),
        ("", False),
        ("TIP: Delete the grey example rows before uploading!", True),
    ]
    ws2.column_dimensions['A'].width = 80
    for row_idx, (text, bold) in enumerate(instructions, 1):
        cell = ws2.cell(row=row_idx, column=1, value=text)
        cell.font = Font(bold=bold, size=10, name="Calibri")
        cell.alignment = Alignment(wrap_text=True)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def create_output_excel(results, selected_fields, custom_fields):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scraped Data"
    all_fields = selected_fields + [f for f in custom_fields if f.strip()]
    headers = ['Part Number', 'Status', 'Source URL']
    for field in all_fields:
        headers.append(field)
        headers.append(f'{field} — Confidence')
    header_fill = PatternFill(start_color="0D1B3E", end_color="0D1B3E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
    conf_fill = PatternFill(start_color="1A3A6B", end_color="1A3A6B", fill_type="solid")
    high_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    medium_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    low_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'), bottom=Side(style='thin', color='E0E0E0')
    )
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = conf_fill if '— Confidence' in header else header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[1].height = 35
    for row_idx, result in enumerate(results, 2):
        for col_idx, header in enumerate(headers, 1):
            value = result.get(header, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=9)
            if '— Confidence' in header and value:
                if '✅ High' in str(value): cell.fill = high_fill
                elif '⚠️ Medium' in str(value): cell.fill = medium_fill
                elif '❌ Low' in str(value): cell.fill = low_fill
        ws.row_dimensions[row_idx].height = 60
    col_widths = {'Part Number': 18, 'Status': 20, 'Source URL': 45}
    for field in all_fields:
        col_widths[field] = 40
        col_widths[f'{field} — Confidence'] = 20
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = col_widths.get(header, 20)
    ws.freeze_panes = 'A2'
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/download-template')
def download_template():
    output = create_template_excel()
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='Pattern_Scraper_Template.xlsx')

@app.route('/scrape', methods=['POST'])
def scrape():
    try:
        file = request.files.get('file')
        selected_fields = json.loads(request.form.get('fields', '[]'))
        custom_fields_raw = request.form.get('custom_fields', '').strip()
        custom_fields = [f.strip() for f in custom_fields_raw.split('\n') if f.strip()] if custom_fields_raw else []
        if not file:
            return jsonify({'error': 'No file uploaded'}), 400
        if not selected_fields and not custom_fields:
            return jsonify({'error': 'Please select at least one field to scrape'}), 400
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        products = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[0]).strip() and 'EXAMPLE' not in str(row[0]).upper():
                products.append({
                    'part_number': str(row[0]).strip(),
                    'brand_website': str(row[1]).strip() if row[1] else '',
                    'product_url': str(row[2]).strip() if len(row) > 2 and row[2] else '',
                })
        if not products:
            return jsonify({'error': 'No products found. Make sure to delete the grey example rows!'}), 400
        results = []
        for product in products:
            result = scrape_product(product['part_number'], product['brand_website'],
                                    product['product_url'], selected_fields, custom_fields)
            results.append(result)
        output = create_output_excel(results, selected_fields, custom_fields)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name='Pattern_Scraped_Data.xlsx')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, port=5001)
